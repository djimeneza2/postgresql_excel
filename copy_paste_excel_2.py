import openpyxl
import datetime

inicio_ejecucion=datetime.datetime.now()
print(inicio_ejecucion)

#####################################################################
def copyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    for i in range(startRow,endRow + 1,1):
        rowSelected = []
        for j in range(startCol,endCol+1,1):
            rowSelected.append(sheet.cell(row = i, column = j).value)
        rangeSelected.append(rowSelected)
    return rangeSelected

def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving,copiedData):
    countRow = 0
    for i in range(startRow,endRow+1,1):
        countCol = 0
        for j in range(startCol,endCol+1,1):            
            sheetReceiving.cell(row = i, column = j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1

####################################################################################
'''

###################             ejecucion           #######################

path_root='/config/workspace/proyecto_oswalk/root_inicio'
path_client='/'+'EDLN'
path_year='/'+'2022'
path_month= '/'+'05_2022'
archive='Modelo_Fact-EDLN_'+'05_2022.xlsx'
final_copy_path=path_root+path_client+path_year+path_month+'/'+archive
wb = openpyxl.load_workbook(final_copy_path,data_only=True)
sheet = wb["Energia"]

path_root_paste='/config/workspace/proyecto_oswalk/root-medicion-sic'
path_year_paste='/'+'2022'
path_month_paste= '/'+'05-2022'
archive_paste='EDLN_M.xlsx'
final_paste_path=path_root_paste+path_year_paste+path_month_paste+'/'+archive_paste
template = openpyxl.load_workbook(final_paste_path) 
temp_sheet = template["medicion"] 

selectedRange = copyRange(1,1,200,3000,sheet) 
pastingRange = pasteRange(1,1,200,3000,temp_sheet,selectedRange) 
template.save(final_paste_path)

##########################################################

path_root='/config/workspace/proyecto_oswalk/root_inicio'
path_client='/'+'EDLN'
path_year='/'+'2022'
path_month= '/'+'05_2022'
archive='Calculo-FTi-EDLN-'+'05_2022.xlsx'
final_copy_path=path_root+path_client+path_year+path_month+'/'+archive
wb = openpyxl.load_workbook(final_copy_path,data_only=True)
sheet = wb["FTi-CMg"]

path_root_paste='/config/workspace/proyecto_oswalk/root-medicion-sic'
path_year_paste='/'+'2022'
path_month_paste= '/'+'05-2022'
archive_paste='EDLN_M.xlsx'
final_paste_path=path_root_paste+path_year_paste+path_month_paste+'/'+archive_paste
template = openpyxl.load_workbook(final_paste_path) 
temp_sheet = template["totalizador"] 

selectedRange = copyRange(1,1,200,3000,sheet) 
pastingRange = pasteRange(1,1,200,3000,temp_sheet,selectedRange) 
template.save(final_paste_path)

print('############### tiempo ejecucion')
final_ejecucion=datetime.datetime.now()
print(final_ejecucion-inicio_ejecucion)

##############################################################################################


###################             ejecucion           #######################

path_root='/config/workspace/proyecto_oswalk/root_inicio'
path_client='/'+'LDS'
path_year='/'+'2022'
path_month= '/'+'05_2022'
archive='Modelo_Fact-LDS_'+'05_2022.xlsx'
final_copy_path=path_root+path_client+path_year+path_month+'/'+archive
wb = openpyxl.load_workbook(final_copy_path,data_only=True)
sheet = wb["Energia"]

path_root_paste='/config/workspace/proyecto_oswalk/root-medicion-sic'
path_year_paste='/'+'2022'
path_month_paste= '/'+'05-2022'
archive_paste='LDSS_M.xlsx'
final_paste_path=path_root_paste+path_year_paste+path_month_paste+'/'+archive_paste
template = openpyxl.load_workbook(final_paste_path) 
temp_sheet = template["Energia"] 

selectedRange = copyRange(1,1,200,3000,sheet) 
pastingRange = pasteRange(1,1,200,3000,temp_sheet,selectedRange) 
template.save(final_paste_path)

##########################################################

path_root='/config/workspace/proyecto_oswalk/root_inicio'
path_client='/'+'LDS'
path_year='/'+'2022'
path_month= '/'+'05_2022'
archive='Calculo-FTi-LDS-'+'05_2022.xlsx'
final_copy_path=path_root+path_client+path_year+path_month+'/'+archive
wb = openpyxl.load_workbook(final_copy_path,data_only=True)
sheet = wb["FTi-CMg"]

path_root_paste='/config/workspace/proyecto_oswalk/root-medicion-sic'
path_year_paste='/'+'2022'
path_month_paste= '/'+'05-2022'
archive_paste='LDSS_M.xlsx'
final_paste_path=path_root_paste+path_year_paste+path_month_paste+'/'+archive_paste
template = openpyxl.load_workbook(final_paste_path) 
temp_sheet = template["FTi-CMg"] 

selectedRange = copyRange(1,1,200,3000,sheet) 
pastingRange = pasteRange(1,1,200,3000,temp_sheet,selectedRange) 
template.save(final_paste_path)

print('############### tiempo ejecucion')
final_ejecucion=datetime.datetime.now()
print(final_ejecucion-inicio_ejecucion)
'''
##############################################################################################

###################             ejecucion           #######################

path_root='/config/workspace/root_inicio'
path_client='/'+'ENOSA'
path_year='/'+'2022'
path_month= '/'+'05_2022'
archive='Modelo_Fact-ENOSA_'+'05_2022.xlsx'
final_copy_path=path_root+path_client+path_year+path_month+'/'+archive
wb = openpyxl.load_workbook(final_copy_path,data_only=True)
sheet = wb["Energia"]

path_root_paste='/config/workspace/root-medicion-sic'
path_year_paste='/'+'2022'
path_month_paste= '/'+'05-2022'
archive_paste='ENOS_M.xlsx'
final_paste_path=path_root_paste+path_year_paste+path_month_paste+'/'+archive_paste
template = openpyxl.load_workbook(final_paste_path) 
temp_sheet = template["Energia"] 

selectedRange = copyRange(1,1,34,2986,sheet) 
pastingRange = pasteRange(1,1,34,2986,temp_sheet,selectedRange) 
template.save(final_paste_path)

##########################################################

path_root='/config/workspace/root_inicio'
path_client='/'+'ENOSA'
path_year='/'+'2022'
path_month= '/'+'05_2022'
archive='Calculo-FTi-ENOSA-'+'05_2022.xlsx'
final_copy_path=path_root+path_client+path_year+path_month+'/'+archive
wb = openpyxl.load_workbook(final_copy_path,data_only=True)
sheet = wb["FTi-CMg"]

path_root_paste='/config/workspace/root-medicion-sic'
path_year_paste='/'+'2022'
path_month_paste= '/'+'05-2022'
archive_paste='ENOS_M.xlsx'
final_paste_path=path_root_paste+path_year_paste+path_month_paste+'/'+archive_paste
template = openpyxl.load_workbook(final_paste_path) 
temp_sheet = template["FTi-CMg"] 
#copyRange(startCol, startRow, endCol, endRow, sheet)
selectedRange = copyRange(1,1,11,2986,sheet) 
pastingRange = pasteRange(1,1,11,2986,temp_sheet,selectedRange) 
template.save(final_paste_path)

##########################################################
path_root='/config/workspace/root_inicio'
path_client='/'+'ENOSA'
path_year='/'+'2022'
path_month= '/'+'05_2022'
archive='Modelo_Fact-ENOSA_'+'05_2022.xlsx'
final_copy_path=path_root+path_client+path_year+path_month+'/'+archive
wb = openpyxl.load_workbook(final_copy_path,data_only=True)
sheet = wb["ReactivaEXC"]

path_root_paste='/config/workspace/root-medicion-sic'
path_year_paste='/'+'2022'
path_month_paste= '/'+'05-2022'
archive_paste='ENOS_M.xlsx'
final_paste_path=path_root_paste+path_year_paste+path_month_paste+'/'+archive_paste
template = openpyxl.load_workbook(final_paste_path) 
temp_sheet = template["ReactivaEXC"] 

selectedRange = copyRange(1,1,34,2986,sheet) 
pastingRange = pasteRange(1,1,34,2986,temp_sheet,selectedRange) 
template.save(final_paste_path)

##########################################################
path_root='/config/workspace/root_inicio'
path_client='/'+'ENOSA'
path_year='/'+'2022'
path_month= '/'+'05_2022'
archive='Modelo_Fact-ENOSA_'+'05_2022.xlsx'
final_copy_path=path_root+path_client+path_year+path_month+'/'+archive
wb = openpyxl.load_workbook(final_copy_path,data_only=True)
sheet = wb["Energia-Libres"]

path_root_paste='/config/workspace/root-medicion-sic'
path_year_paste='/'+'2022'
path_month_paste= '/'+'05-2022'
archive_paste='ENOS_M.xlsx'
final_paste_path=path_root_paste+path_year_paste+path_month_paste+'/'+archive_paste
template = openpyxl.load_workbook(final_paste_path) 
temp_sheet = template["Energia-Libres"] 

selectedRange = copyRange(1,1,34,2986,sheet) 
pastingRange = pasteRange(1,1,34,2986,temp_sheet,selectedRange) 
template.save(final_paste_path)

##########################################################
path_root='/config/workspace/root_inicio'
path_client='/'+'ENOSA'
path_year='/'+'2022'
path_month= '/'+'05_2022'
archive='Modelo_Fact-ENOSA_'+'05_2022.xlsx'
final_copy_path=path_root+path_client+path_year+path_month+'/'+archive
wb = openpyxl.load_workbook(final_copy_path,data_only=True)
sheet = wb["Energia-stevia"]

path_root_paste='/config/workspace/root-medicion-sic'
path_year_paste='/'+'2022'
path_month_paste= '/'+'05-2022'
archive_paste='ENOS_M.xlsx'
final_paste_path=path_root_paste+path_year_paste+path_month_paste+'/'+archive_paste
template = openpyxl.load_workbook(final_paste_path) 
temp_sheet = template["Energia-stevia"] 

selectedRange = copyRange(1,1,34,2986,sheet) 
pastingRange = pasteRange(1,1,34,2986,temp_sheet,selectedRange) 
template.save(final_paste_path)

##########################################################
path_root='/config/workspace/root_inicio'
path_client='/'+'ENOSA'
path_year='/'+'2022'
path_month= '/'+'05_2022'
archive='Modelo_Fact-ENOSA_'+'05_2022.xlsx'
final_copy_path=path_root+path_client+path_year+path_month+'/'+archive
wb = openpyxl.load_workbook(final_copy_path,data_only=True)
sheet = wb["Energia-Libres-1"]

path_root_paste='/config/workspace/root-medicion-sic'
path_year_paste='/'+'2022'
path_month_paste= '/'+'05-2022'
archive_paste='ENOS_M.xlsx'
final_paste_path=path_root_paste+path_year_paste+path_month_paste+'/'+archive_paste
template = openpyxl.load_workbook(final_paste_path) 
temp_sheet = template["Energia-Libres-1"] 

selectedRange = copyRange(1,1,34,2986,sheet) 
pastingRange = pasteRange(1,1,34,2986,temp_sheet,selectedRange) 
template.save(final_paste_path)

##########################################################
path_root='/config/workspace/root_inicio'
path_client='/'+'ENOSA'
path_year='/'+'2022'
path_month= '/'+'05_2022'
archive='Modelo_Fact-ENOSA_'+'05_2022.xlsx'
final_copy_path=path_root+path_client+path_year+path_month+'/'+archive
wb = openpyxl.load_workbook(final_copy_path,data_only=True)
sheet = wb["Energia-Libres-2"]

path_root_paste='/config/workspace/root-medicion-sic'
path_year_paste='/'+'2022'
path_month_paste= '/'+'05-2022'
archive_paste='ENOS_M.xlsx'
final_paste_path=path_root_paste+path_year_paste+path_month_paste+'/'+archive_paste
template = openpyxl.load_workbook(final_paste_path) 
temp_sheet = template["Energia-Libres-2"] 

selectedRange = copyRange(1,1,34,2986,sheet) 
pastingRange = pasteRange(1,1,34,2986,temp_sheet,selectedRange) 
template.save(final_paste_path)







print('############### tiempo ejecucion')
final_ejecucion=datetime.datetime.now()
print(final_ejecucion-inicio_ejecucion)