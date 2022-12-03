import os
import openpyxl
import pandas as pd
import numpy as np
import datetime
import psycopg2
from sqlalchemy import create_engine

from A_query_postgresql import *



class check_archives_in_path():

    def __init__(self,final_path):

        self.final_path = final_path
    
    def read_archives_in_path(self):

        os.chdir(self.final_path)

        self.array_of_archives=[]

        for root, dirs, files in os.walk('.', topdown = False):

            for name in files:

                self.array_of_archives.append(name)

        return self.array_of_archives



class copy_excel_data():

    def __init__(self, final_path, workbook, sheet, startRow, startCol, endRow, endCol):

        self.final_path = final_path

        self.workbook = workbook

        self.sheet = sheet

        self.startRow = startRow

        self.startCol = startCol

        self.endRow = endRow

        self.endCol = endCol
    
    def activate_workbook_to_copy(self):

        wb = openpyxl.load_workbook( self.final_path, data_only=True )

        wb_sheet_copy = wb[ self.sheet ]

        self.rangeSelected = []

        for i in range( self.startRow, self.endRow + 1, 1):

            rowSelected = []

            for j in range( self.startCol, self.endCol+1, 1):

                rowSelected.append( wb_sheet_copy.cell(row = i, column = j).value)

            self.rangeSelected.append(rowSelected)

        return self.rangeSelected



class paste_excel_data():

    def __init__(self, final_path, workbook, sheet, startRow, startCol, endRow, endCol, copiedData):

        self.final_path = final_path

        self.workbook = workbook

        self.sheet = sheet

        self.startRow = startRow

        self.startCol = startCol

        self.endRow = endRow

        self.endCol = endCol   

        self.copiedData = copiedData

    def activate_workbook_to_paste(self):

        wb = openpyxl.load_workbook( self.final_path )

        wb_sheet_paste = wb[ self.sheet ]
        
        countRow = 0
        
        for i in range(self.startRow,self.endRow+1,1):
            
            countCol = 0
            
            for j in range(self.startCol,self.endCol+1,1):            
                
                wb_sheet_paste.cell(row = i, column = j).value = self.copiedData[countRow][countCol]
                
                countCol += 1
            
            countRow += 1

        wb.save( self.final_path )

        return 0



class create_timestamp_for_dataframe():

    def __init__(self,year,month):

        self.year=year

        self.month=month

    def create_timestamp_array(self):

        if self.month in [1,3,5,7,8,10,12]:

            self.days=31

        elif self.month in [4,6,9,11]:

            self.days=30

        elif self.month in [2]:

            if self.year%4 == 0 and (self.year%100 != 0 or self.year%400==0):

                self.days=29

            else:

                self.days=28

        self.array_timestamp = []

        x=datetime.datetime(self.year,self.month,1,0,0)

        for i in range(self.days*24*4):

            x+=datetime.timedelta(minutes = 15)

            self.array_timestamp.append([x])

        return self.array_timestamp 

    def create_zero_arrays(self):

        self.array_zeros_df = np.zeros( [self.days*24*4, 1] )

        return self.array_zeros_df


    def create_ones_arrays(self):

        self.array_ones_df = np.ones( [self.days*24*4, 1] )

        return self.array_ones_df



#######################################################

inicio_ejecucion=datetime.datetime.now()
print(inicio_ejecucion)

search_path_root='/config/workspace/root_inicio'
search_path_client='/'+'ENOSA'
search_path_year='/'+'2022'
search_path_month= '/'+'05_2022'
search_final_path=search_path_root+search_path_client+search_path_year+search_path_month
search_final_path_data=search_final_path+'/'+'intefase_postgresql.xlsx'
all_client_mesures=check_archives_in_path(search_final_path+'/'+'052022_ENG_DC').read_archives_in_path()
data_postgresql=pd.read_excel(search_final_path_data,sheet_name="nombre_data")
tabla_postgresql=pd.read_excel(search_final_path_data,sheet_name="tablas_postgres")
search_excel_name = data_postgresql['nombre_medicion_enviada'].isin(all_client_mesures)

par_postgresql=[]
for j in data_postgresql.index:
    par=[data_postgresql['nombre_medicion_enviada'][j],data_postgresql['encontrado_numero_postgres'][j],data_postgresql['mediciones_tabla_postgres'][j]]
    par_postgresql.append(par)

timestamps=create_timestamp_for_dataframe(2022,5)
array_timestamp=timestamps.create_timestamp_array()
array_kvar_i=timestamps.create_zero_arrays()
array_kvar_c=timestamps.create_zero_arrays()


for client,numero,tabla in par_postgresql:

    copy_path_root='/config/workspace/root_inicio'
    copy_path_client='/'+'ENOSA'
    copy_path_year='/'+'2022'
    copy_path_month= '/'+'05_2022'+'/'+'052022_ENG_DC'
    copy_archive=client
    copy_sheet='Compra'
    copy_final_path=copy_path_root+copy_path_client+ copy_path_year+ copy_path_month+ '/'+copy_archive
    print(copy_final_path)

    kwh=copy_excel_data(copy_final_path,copy_archive,copy_sheet,5,3,2980,3)
    kwh_copied = kwh.activate_workbook_to_copy()

    kw=copy_excel_data(copy_final_path,copy_archive,copy_sheet,5,2,2980,2)
    kw_copied = kw.activate_workbook_to_copy()

    array_id_facturacion = np.dot(timestamps.create_ones_arrays(),numero)

    data_dataframe = np.concatenate((kwh_copied,
                                                array_kvar_i,
                                                array_kvar_c,
                                                kw_copied,
                                                kwh_copied,
                                                array_id_facturacion,
                                                array_timestamp),axis=1)

    dataframe_prueba=pd.DataFrame(data_dataframe ,
                                                columns=['kwh',
                                                        'kvar_i',
                                                        'kvar_c',
                                                        'kw',
                                                        'kwh_i',
                                                        'id_facturacion',
                                                        'periodo'])
    

    
    records_to_insert=[]

    for ro in dataframe_prueba.index: 

        tuple_prueba=()

        for col in dataframe_prueba.columns: 

            tuple_prueba+=(dataframe_prueba.loc[ro,col],)

        records_to_insert.append(tuple_prueba)

    data_base_conection("admin",
                    "secret",
                    "172.25.0.1",
                    "5432",
                    "mediciones_cliente"
                    ).execute_query_insert_many(tabla,records_to_insert)

    dataframe_prueba.to_csv(copy_path_root+
                        copy_path_client+ 
                        copy_path_year+ '/'+'05_2022'+'/revision_libres/'+str(numero)+'.csv')

print('############### tiempo ejecucion')
final_ejecucion=datetime.datetime.now()
print(final_ejecucion-inicio_ejecucion)

###################################################