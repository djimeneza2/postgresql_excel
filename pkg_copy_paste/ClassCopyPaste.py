
'''
import os
import openpyxl
import pandas as pd
import numpy as np
import datetime
import psycopg2
from sqlalchemy import create_engine
import xlwings as xw
import gc
'''
import pkg_copy_paste

class copy_excel_data():

    def __init__(self, 
                 final_path, 
                 workbook, 
                 sheet, 
                 startRow, 
                 startCol, 
                 endRow, 
                 endCol):

        self.final_path = final_path
        self.workbook = workbook
        self.sheet = sheet
        self.startRow = startRow
        self.startCol = startCol
        self.endRow = endRow
        self.endCol = endCol
    
    def activate_workbook_to_copy(self):

        excel_app = pkg_copy_paste.xw.App(visible=False)
        excel_book = excel_app.books.open(self.final_path+self.workbook,update_links=False)
        excel_book.save()
        excel_book.close()
        excel_app.quit()

        wb = pkg_copy_paste.openpyxl.load_workbook( self.final_path+self.workbook,data_only=True,keep_vba=False)
        #wb = xw.Book( self.final_path+self.workbook)

        wb.active

        wb_sheet_copy = wb.get_sheet_by_name(self.sheet)

        self.rangeSelected = []

        for i in range( self.startRow, self.endRow + 1, 1):

            rowSelected = []

            for j in range( self.startCol, self.endCol+1, 1):

                data=wb_sheet_copy.cell(row = i, column = j).value
 
                #print(data)

                rowSelected.append(data)

            self.rangeSelected.append(rowSelected)

        pkg_copy_paste.gc.collect()

        print('copied data from '+self.final_path+self.workbook+' sheet '+self.sheet+' complete\n')
        #print(self.rangeSelected)

        return self.rangeSelected
    
class copy_excel_data_2(copy_excel_data):

    def __init__(self,
                 final_path, 
                 workbook, 
                 sheet, 
                 startRow, 
                 startCol, 
                 endRow, 
                 endCol):

        super().__init__(final_path, 
                 workbook, 
                 sheet, 
                 startRow, 
                 startCol, 
                 endRow, 
                 endCol)

    def activate_workbook_to_copy(self):

        app = pkg_copy_paste.xw.App(add_book=False)
        app.display_alerts = False

        #wb = xw.Book( self.final_path+self.workbook)
        wb =app.books.open(self.final_path+self.workbook, 
                           update_links=False, 
                           read_only=True, 
                           ignore_read_only_recommended=True)

        wb_sheet_copy = wb.sheets[self.sheet]

        copy_range = wb_sheet_copy.range((self.startRow,self.startCol),(self.endRow,self.endCol))

        self.rangeSelected=[]
        for i in list(copy_range):
            self.rangeSelected.append(i.value)

        print('copied data complete')
                
        print(self.rangeSelected)

        wb.close()
        app.kill()

        return self.rangeSelected
    
class paste_excel_data():

    def __init__(self, 
                 final_path, 
                 workbook, 
                 sheet, 
                 startRow, 
                 startCol, 
                 endRow, 
                 endCol, 
                 copiedData):

        self.final_path = final_path
        self.workbook = workbook
        self.sheet = sheet
        self.startRow = startRow
        self.startCol = startCol
        self.endRow = endRow
        self.endCol = endCol   
        self.copiedData = copiedData

    def activate_workbook_to_paste(self):

        wb = pkg_copy_paste.openpyxl.load_workbook( self.final_path+self.workbook)
        #wb = xw.Book( self.final_path+self.workbook)

        wb_sheet_paste = wb.get_sheet_by_name(self.sheet)
        
        countRow = 0
        
        for i in range(self.startRow,self.endRow+1,1):
            
            countCol = 0
            
            for j in range(self.startCol,self.endCol+1,1):            
                
                wb_sheet_paste.cell(row = i, column = j).value = self.copiedData[countRow][countCol]
                
                countCol += 1
            
            countRow += 1
        
        pkg_copy_paste.gc.collect()

        wb.save( self.final_path+self.workbook)

        return print('paste data to '+ self.final_path+self.workbook + ' sheet '+self.sheet +' complete\n')
    
class paste_excel_data_2(paste_excel_data):

    def __init__(self, 
                 final_path, 
                 workbook, 
                 sheet, 
                 startRow, 
                 startCol, 
                 endRow, 
                 endCol, 
                 copiedData):
        
        super().__init__(final_path, 
                 workbook, 
                 sheet, 
                 startRow, 
                 startCol, 
                 endRow, 
                 endCol, 
                 copiedData)

    def activate_workbook_to_paste(self):

        app = pkg_copy_paste.xw.App(add_book=False)
        app.display_alerts = False

        #wb = xw.Book( self.final_path+self.workbook)
        wb =app.books.open(self.final_path+self.workbook, 
                           update_links=False, 
                           ignore_read_only_recommended=True)

        wb_sheet_paste = wb.sheets[self.sheet]
        
        wb_sheet_paste.range((self.startRow,self.startCol),(self.endRow,self.endCol)).value=self.copiedData

        wb.save(self.final_path+self.workbook)
        wb.close()
        app.kill()

        return print('paste data complete')
