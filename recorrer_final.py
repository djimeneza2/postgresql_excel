import os
import openpyxl
import pandas as pd
import numpy as np
import datetime



class check_archives_in_path():

    def __init__(self,final_path):

        self.final_path = final_path
    
    def read_archives_in_path(self):

        os.chdir(self.final_path)

        self.array_of_archives=[]

        for root, dirs, files in os.walk('.', topdown = False):

            for name in files:

                self.array_of_archives.append(name)

        return self.array_of_archives



class copy_excel_data():

    def __init__(self, final_path, workbook, sheet, startRow, startCol, endRow, endCol):

        self.final_path = final_path

        self.workbook = workbook

        self.sheet = sheet

        self.startRow = startRow

        self.startCol = startCol

        self.endRow = endRow

        self.endCol = endCol
    
    def activate_workbook_to_copy(self):

        wb = openpyxl.load_workbook( self.final_path, data_only=True )

        wb_sheet_copy = wb[ self.sheet ]

        self.rangeSelected = []

        for i in range( self.startRow, self.endRow + 1, 1):

            rowSelected = []

            for j in range( self.startCol, self.endCol+1, 1):

                rowSelected.append( wb_sheet_copy.cell(row = i, column = j).value)

            self.rangeSelected.append(rowSelected)

        return self.rangeSelected



class paste_excel_data():

    def __init__(self, final_path, workbook, sheet, startRow, startCol, endRow, endCol, copiedData):

        self.final_path = final_path

        self.workbook = workbook

        self.sheet = sheet

        self.startRow = startRow

        self.startCol = startCol

        self.endRow = endRow

        self.endCol = endCol   

        self.copiedData = copiedData

    def activate_workbook_to_paste(self):

        wb = openpyxl.load_workbook( self.final_path )

        wb_sheet_paste = wb[ self.sheet ]
        
        countRow = 0
        
        for i in range(self.startRow,self.endRow+1,1):
            
            countCol = 0
            
            for j in range(self.startCol,self.endCol+1,1):            
                
                wb_sheet_paste.cell(row = i, column = j).value = self.copiedData[countRow][countCol]
                
                countCol += 1
            
            countRow += 1

        wb.save( self.final_path )

        return 0



#######################################################

inicio_ejecucion=datetime.datetime.now()
print(inicio_ejecucion)

search_path_root='/config/workspace/root_inicio'
search_path_client='/'+'ENOSA'
search_path_year='/'+'2022'
search_path_month= '/'+'05_2022'
search_path_med='/'+'052022_ENG_DC'
search_archives_med='/'+'mediciones_libres.xlsx'
search_final_path=search_path_root+search_path_client+search_path_year+search_path_month
search_final_path_data=search_final_path+search_archives_med
all_client_mesures=check_archives_in_path(search_final_path+search_path_med).read_archives_in_path()
data_postgresql=pd.read_excel(search_final_path_data,sheet_name="nombre_data")
tabla_postgresql=pd.read_excel(search_final_path_data,sheet_name="tablas_postgres")
search_excel_name = data_postgresql['nombre_medicion_enviada'].isin(all_client_mesures)
print(search_excel_name)

par_postgresql=[]
for j in data_postgresql.index:
    par=[data_postgresql['nombre_medicion_enviada'][j],data_postgresql['encontrado_numero_postgres'][j]]
    par_postgresql.append(par)

dict_dataframe_medicion_libres={}
for client,numero in par_postgresql:
    copy_path_root='/config/workspace/root_inicio'
    copy_path_client='/'+'ENOSA'
    copy_path_year='/'+'2022'
    copy_path_month= '/'+'05_2022'
    copy_path_med='/'+'052022_ENG_DC'
    copy_archive=client
    copy_sheet='Compra'
    copy_final_path=copy_path_root+copy_path_client+ copy_path_year+ copy_path_month+copy_path_med+'/'+copy_archive
    create_archive_csv='/'+'revision_libres'
    print(copy_final_path)

    kwh=copy_excel_data(copy_final_path,copy_archive,copy_sheet,5,3,2980,3)
    kwh_copied = kwh.activate_workbook_to_copy()

    kw=copy_excel_data(copy_final_path,copy_archive,copy_sheet,5,2,2980,2)
    kw_copied = kw.activate_workbook_to_copy()

    dataframe_prueba=pd.DataFrame(np.concatenate((kwh_copied,kw_copied),axis=1),columns=['kwh','kw'])
    dataframe_prueba.to_csv(copy_path_root+copy_path_client+ copy_path_year+ '/'+copy_path_month+'/'+create_archive_csv+'/'+str(numero)+'.csv')
    dict_dataframe_medicion_libres[numero]=dataframe_prueba

final_ejecucion=datetime.datetime.now()
print(final_ejecucion-inicio_ejecucion)

###################################################