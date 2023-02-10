import os
import openpyxl
import pandas as pd
import numpy as np
import datetime
import psycopg2
from sqlalchemy import create_engine



class check_archives_in_path():

    def __init__(self,final_path):

        self.final_path = final_path
    
    def read_archives_in_path(self):

        os.chdir(self.final_path)

        self.array_of_archives=[]

        for root, dirs, files in os.walk('.', topdown = False):

            for name in files:

                self.array_of_archives.append(name)

        return self.array_of_archives



class copy_excel_data():

    def __init__(self, final_path, workbook, sheet, startRow, startCol, endRow, endCol):

        self.final_path = final_path

        self.workbook = workbook

        self.sheet = sheet

        self.startRow = startRow

        self.startCol = startCol

        self.endRow = endRow

        self.endCol = endCol
    
    def activate_workbook_to_copy(self):

        wb = openpyxl.load_workbook( self.final_path, data_only=True )

        wb_sheet_copy = wb[ self.sheet ]

        self.rangeSelected = []

        for i in range( self.startRow, self.endRow + 1, 1):

            rowSelected = []

            for j in range( self.startCol, self.endCol+1, 1):

                if wb_sheet_copy.cell(row = i, column = j).value != None:

                    rowSelected.append( wb_sheet_copy.cell(row = i, column = j).value)

                else:

                    rowSelected.append( 0 )

            self.rangeSelected.append(rowSelected)

        return self.rangeSelected



class paste_excel_data():

    def __init__(self, final_path, workbook, sheet, startRow, startCol, endRow, endCol, copiedData):

        self.final_path = final_path

        self.workbook = workbook

        self.sheet = sheet

        self.startRow = startRow

        self.startCol = startCol

        self.endRow = endRow

        self.endCol = endCol   

        self.copiedData = copiedData

    def activate_workbook_to_paste(self):

        wb = openpyxl.load_workbook( self.final_path )

        wb_sheet_paste = wb[ self.sheet ]
        
        countRow = 0
        
        for i in range(self.startRow,self.endRow+1,1):
            
            countCol = 0
            
            for j in range(self.startCol,self.endCol+1,1):            
                
                wb_sheet_paste.cell(row = i, column = j).value = self.copiedData[countRow][countCol]
                
                countCol += 1
            
            countRow += 1

        wb.save( self.final_path )

        return 0



class create_timestamp_for_dataframe():

    def __init__(self,year,month):

        self.year=year

        self.month=month

    def create_timestamp_array(self):

        if self.month in [1,3,5,7,8,10,12]:

            self.days=31

        elif self.month in [4,6,9,11]:

            self.days=30

        elif self.month in [2]:

            if self.year%4 == 0 and (self.year%100 != 0 or self.year%400==0):

                self.days=29

            else:

                self.days=28

        self.array_timestamp = []

        x=datetime.datetime(self.year,self.month,1,0,0)

        for i in range(self.days*24*4):

            x+=datetime.timedelta(minutes = 15)

            self.array_timestamp.append([x])

        return self.array_timestamp 

    def create_zero_arrays(self):

        self.array_zeros_df = np.zeros( [self.days*24*4, 1] )

        return self.array_zeros_df


    def create_ones_arrays(self):

        self.array_ones_df = np.ones( [self.days*24*4, 1] )

        return self.array_ones_df

    def create_range_arrays(self):

        self.array_range_df = []
        for i in range(1,self.days*24*4+1):
            self.array_range_df.append([i])

        return self.array_range_df